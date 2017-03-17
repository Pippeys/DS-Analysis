import openpyxl, pprint
import math

wb_BART = openpyxl.load_workbook('Bart Stations.xlsx', data_only=True)
wb_Properties = openpyxl.load_workbook('Bart Distance.xlsx', data_only=True)
BART_sheet = wb_BART.get_sheet_by_name('Stations')
prop_sheet = wb_Properties.get_sheet_by_name()

#Lists of properties and BART Stations
prop_lat = []
prop_long = []
bart_lat = []
bart_long = []
d_lat = []
d_long = []

def Bart_Coor():
    for row in range(2,BART_sheet.max_row + 1):
        b_lat = BART_sheet['B' + str(row)].value
        b_long = BART_sheet['C' + str(row)].value
        bart_lat.append(b_lat)
        bart_long.append(b_long)

def Prop_Coor():
    for row in range(2,prop_sheet.max_row + 1):
        p_lat = prop_sheet['D' + str(row)].value
        p_long = prop_sheet['C' + str(row)].value
        prop_lat.append(p_lat)
        prop_long.append(p_long)

def Diff_Coor():
    for row in range(2,prop_sheet.max_row + 1):
        for i in range(2, BART_sheet.max_row + 1):
            dlat = prop_lat(row) - bart_lat(i)
            d_lat.append(dlat)

def Main():
    Bart_Coor()
    Prop_Coor()
    Diff_Coor()
    print(d_lat)
